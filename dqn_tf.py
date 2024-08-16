# string

input = 'Hello how are you'
words = input.split(' ')

words.reverse()
reverse_sentence = ' '.join(words)

print(reverse_sentence)

# Dictionary

theBoard = {'7': ' ', '8': ' ', '9': ' ',
            '4': ' ', '5': ' ', '6': ' ',
            '1': ' ', '2': ' ', '3': ' '}

def printBoard(board):
    print(board['7'] + '|' + board['8'] + '|' + board['9'])
    print('-+-+-')
    print(board['4'] + '|' + board['5'] + '|' + board['6'])
    print('-+-+-')
    print(board['1'] + '|' + board['2'] + '|' + board['3'])

def game():
    turn = 'X'
    count = 0

    for i in range(10):
        printBoard(theBoard)
        print("It's your turn," + turn + ".Move to which place?")

        move = input()

        if theBoard[move] == ' ':
            theBoard[move] = turn
            count += 1
        else:
            print("That place is already filled.\nMove to which place?")
            continue

        if count >= 5:
            if theBoard['7'] == theBoard['8'] == theBoard['9'] != ' ':  # across the top
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break
            elif theBoard['4'] == theBoard['5'] == theBoard['6'] != ' ':  # across the middle
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break
            elif theBoard['1'] == theBoard['2'] == theBoard['3'] != ' ':  # across the bottom
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break
            elif theBoard['1'] == theBoard['4'] == theBoard['7'] != ' ':  # down the left side
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break
            elif theBoard['2'] == theBoard['5'] == theBoard['8'] != ' ':  # down the middle
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break
            elif theBoard['3'] == theBoard['6'] == theBoard['9'] != ' ':  # down the right side
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break
            elif theBoard['7'] == theBoard['5'] == theBoard['3'] != ' ':  # diagonal
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break
            elif theBoard['1'] == theBoard['5'] == theBoard['9'] != ' ':  # diagonal
                printBoard(theBoard)
                print("\nGame Over.\n")
                print(" **** " + turn + " won. ****")
                break

        if count == 9:
            print("\nGame Over.\n")
            print("It's a Tie!!")

        if turn == 'X':
            turn = 'O'
        else:
            turn = 'X'


if __name__ == "__main__":
    game()

# RegularExpression

import re

def isEmail(email):
    is_email = re.search(r'^[A-z0-9]+[\.-]?[A-z0-9]+@\w+\.\w{2,3}$', email)

    if is_email:
        print(f'the {email} is a valid Email')

    else:
        print(f'the {email} is not a valid Email')


print('Is hsoub.academy@gmail.com a email?')
isEmail('hsoub.academy@gmail.com')
print('Is hsoub.academy@gmail a email?')
isEmail('hsoub.academy@gmail')

# ZipFiles

import shutil, os, re
from pathlib import Path

datePattern = "^(.*?)((0|1)?\d)-((0|1|2|3)?\d)-((19|20)\d\d)(.*?)$"

for filename in os.listdir(Path.home() / Path('Desktop','files')):
    search = re.search(datePattern, filename)

    if search == None:
        continue

    beforePart = search.group(1)
    monthPart = search.group(2)
    dayPart = search.group(4)
    yearPart = search.group(6)
    afterPart = search.group(8)

    newFilename = beforePart + dayPart + '-' + monthPart + '-' + yearPart + afterPart
    print(f'Renaming "{filename}" to "{newFilename}"')

    oldName = Path.home() / Path('Desktop', 'files') / filename
    newName = Path.home() / Path('Desktop', 'files') / newFilename
    shutil.move(oldName, newName)

# Excel

import openpyxl, sys
from pathlib import Path
from openpyxl.styles import Font

if len(sys.argv) == 2:

    try:
        number = int(sys.argv[1])

    except Exception as e:
        print(e)

    excelFile = openpyxl.Workbook()
    sheet = excelFile.active

    for y in range(number + 1):
        for x in range(number + 1):

            # Check if in header row or column.
            isHeader = False

            if x == 0 and y == 0:
                isHeader = True
                n = ''

            elif x == 0:
                isHeader = True
                n = y

            elif y == 0:
                isHeader = True
                n = x

            else:
                n = x * y

            cell = sheet.cell(row=y + 1, column=x + 1)

            if isHeader:
                cell.font = Font(bold=True)

            cell.value = n

    savedFile = str(Path.home()/ Path('Desktop') / 'multiplication_table_') + str(number) + '.xlsx'

    excelFile.save(savedFile)

    print('Saved as ' + savedFile)

else:
    print('Please enter only two arguments')

# Google

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re

scopes = [
'https://www.googleapis.com/auth/spreadsheets',
'https://www.googleapis.com/auth/drive',
]

credentials = ServiceAccountCredentials.from_json_keyfile_name("keys.json", scopes)
file = gspread.authorize(credentials)

sheet = file.open("example").sheet1

# find
cell = sheet.findall("Hadi")
print(cell)
#print("Found something at Row: %s and Column: %s" % (cell.row, cell.col))

employess = re.compile(r'(Anas|Sara)')
cell = sheet.findall(employess)
print(cell)

# clear
sheet.batch_clear(["A8:C8"])

sheet.clear()

from docx2pdf import convert
from pathlib import Path

docx_file = Path.home() /Path('Onedrive', 'Desktop', 'word files', 'academy_1.docx')
pdf_file = Path.home() / Path('Onedrive', 'Desktop', 'academy_1_converted.pdf')
multi_file = Path.home() / Path('Onedrive', 'Desktop', 'word files/')

convert(multi_file)

# Json

import requests, json

# section 3
api_key = "e59574bf6b887f0e08af237fa247ff89"

# section 1
base_url = "http://api.openweathermap.org/data/2.5/weather?"

# section 2
city_name = input("Enter city name : ")

# complete url address
complete_url = base_url + "q=" + city_name + "&units=metric" + "&appid=" + api_key

# response object
response = requests.get(complete_url)
#print(response.text)

x = json.loads(response.text)

if x["cod"] != "404":

    y = x["main"]

    current_temperature = y["temp"]
    current_pressure = y["pressure"]
    current_humidity = y["humidity"]

    z = x["weather"]

    weather_description = z[0]["description"]

    print(" Temperature (in metric unit) = " +
          str(current_temperature) +
          "\n atmospheric pressure (in metric unit) = " +
          str(current_pressure) +
          "\n humidity (in metric) = " +
          str(current_humidity) +
          "\n description = " +
          str(weather_description))

else:
    print(" City Not Found ")

# CSV

import csv, os
from pathlib import Path

os.makedirs(Path.home() / Path('Desktop', 'CSV files'), exist_ok=True)

for csvFilename in os.listdir(Path.home() / Path('Desktop', 'CSV files')):
     if not csvFilename.endswith('.csv'):
         continue

     print('Removing header from ' + csvFilename + '...')
     csvRows = []
     csvFileObj = open(Path.home() / Path('Desktop', 'CSV files', csvFilename))
     readerObj = csv.reader(csvFileObj)

     for row in readerObj:
         if readerObj.line_num == 1:
             continue  # تخطى السطر الأول
         csvRows.append(row)
     csvFileObj.close()

     csvFileObj = open(Path.home() / Path('Desktop', 'CSV files', csvFilename), 'w', newline='')
     csvWriter = csv.writer(csvFileObj)
     for row in csvRows:
         csvWriter.writerow(row)
     csvFileObj.close()

# Webscrabing

import bs4
import requests
import csv
from pathlib import Path

res = requests.get('https://en.wikipedia.org/wiki/List_of_languages_by_number_of_native_speakers')
#print(res.text)
soup = bs4.BeautifulSoup(res.text, 'html.parser')

# select table
table_soup = soup.find_all("table")
filtered_table = [table for table in table_soup if table.caption is not None]
#print(filtered_table)

required_table = None

for table in filtered_table:
    if str(table.caption.text).strip() == 'Top languages by population per Nationalencyklopedin':
        required_table = table
        break

#print(required_table)

rows = required_table.find_all('tr')
#print(rows)
headers = [ head.text.replace('\n', '') for head in rows[0].find_all('th') ]
#print(headers)

data = []

for row_data in rows:
    value = row_data.find_all('td')
    value_text = [db.text.strip() for db in value]

    if len(value_text) == 0:
        continue

    data.append(value_text)

# write to CSV
file = open(Path.home() / Path('Desktop', 'wikiPedia.csv'), 'w', newline='')
writer = csv.writer(file)
writer.writerow(headers)
writer.writerows(data)
file.close()

# DataBase

import sqlite3
from pathlib import Path

message = """
"a" => Add New Task
"d" => Delete A Task
"s" => Show All Task
"u" => Update A Task
"q" => Quit The App
Please choose an option:
"""

user_input = input(message).strip().lower()

# Command List
commands_list = ["a", "d", "s", "u", "q"]

user_id = 2

try:
  # Connect to DB
  sqliteConnection = sqlite3.connect(Path.home() / Path('Desktop', 'todoAPP.db'))
  crsr = sqliteConnection.cursor()

except:
  print("connection error")

finally:
    if (sqliteConnection):
        # Create Table
        sql_command = """CREATE TABLE if not exists tasks ( 
        user_id INTEGER, 
        task_name VARCHAR(20), 
        description TEXT)"""
        crsr.execute(sql_command)


        def show_tasks():
            crsr.execute(f"SELECT * FROM tasks WHERE user_id = '{user_id}'")

            results = crsr.fetchall()

            print(f"You have {len(results)} tasks")

            if len(results) > 0:
                for task in results:
                    print(f"Task Name: {task[1]} AND", end=" ")
                    print(f"Task Description: {task[2]}")

            sqliteConnection.commit()


        def add_task():
            task_name = input("Write Task Name: ").strip()
            des = input("Write The Task Description: ").strip()

            crsr.execute(f"INSERT INTO tasks (user_id, task_name, description) VALUES ('{user_id}', '{task_name}', '{des}')")

            sqliteConnection.commit()


        def delete_task():
            task_name = input("Write The Task Name You Want To Delete: ").strip()

            crsr.execute(f"DELETE FROM tasks where task_name='{task_name}' and user_id='{user_id}'")
            sqliteConnection.commit()


        def update_task():
            task_name = input("Write The Name Of The Task You Want To Modify: ").strip()
            crsr.execute(f"SELECT * FROM tasks WHERE task_name = '{task_name}' AND user_id='{user_id}'")
            results = crsr.fetchall()

            if not results:
                print("There is no task with this name")

            else:
                des = input("Write The New Task Description: ").strip()
                crsr.execute(
                    f"UPDATE tasks SET description = '{des}' WHERE task_name = '{task_name}' AND user_id = '{user_id}'")

                sqliteConnection.commit()

                print("The task has been successfully modified")


        def end_app():
            print("Program closed")
            exit()


        if user_input in commands_list:

            if user_input == "s":
                show_tasks()

            elif user_input == "a":
                add_task()

            elif user_input == "d":
                delete_task()

            elif user_input == "u":
                update_task()

            else:
                end_app()

        else:
            print("Sorry This Command Is Not Found")

    sqliteConnection.close()

# Email

import openpyxl, smtplib
from pathlib import Path

# connect with Excelfile
file = openpyxl.load_workbook(Path.home() / Path('Onedrive', 'Desktop', 'members.xlsx'))
sheets = file.sheetnames
sheet = file['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
print(latestMonth)

# find unpaid members
unpaidMembers = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Send Email
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.starttls()
sender_email = "academyhsoub1@gmail.com"
password = input(str("Please enter your password: "))
smtpObj.login(sender_email, password)

for name, email in unpaidMembers.items():
    body = """Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues 
    for %s.Please make this payment as soon as possible.Thank you!'""" %(latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail(sender_email, email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()